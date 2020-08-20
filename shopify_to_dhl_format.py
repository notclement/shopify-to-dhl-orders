"""
Created by: Clement

Shopify orders to DHL shipping format

Will be able to let me export orders via Shopify and then convert it into the
format that DHL portal can take in directly
https://ecommerceportal.dhl.com/Portal/pages/login/userlogin.xhtml

STEPS:
1. Logistics team will just export the Shopify orders as csv
2. Run the script with the shopify exported order list
3. DHL formatted csv will be in an output folder
(./output/dhl_<export_dhl_format_YYYY-MM-DD_HHMMSS>.xlsx)
"""

import csv
import sys
from datetime import datetime
from shutil import copyfile

import openpyxl
from openpyxl.utils.cell import column_index_from_string as cifs

from resources.account_details import *
from resources.dhl_index_mapping import *
from resources.shipping_service import *
from resources.shopify_index_mapping import *
from resources.strap_stats import *
from resources.incoterm import dict_incoterm

DHL_DELIMITER = '\t'
SHOPIFY_DELIMITER = ','
DHL_TEMPLATE_FILE = './resources/dhl_csv_headers.xlsx'
EXPORT_NAME = datetime.now().strftime('export_dhl_format_%Y-%m-%d_%H%M%S.xlsx')
EXPORT_PATH_NAME = './site/output/{}'.format(EXPORT_NAME)


def get_shopify_csv_by_line(path):
    """opens and read the shopify csv and yield line by line"""
    with open(path) as csv_file:
        curr_line = 0
        csv_reader = csv.reader(csv_file, delimiter=SHOPIFY_DELIMITER)
        for row in csv_reader:
            # only yield results if its not the header
            if curr_line != 0:
                yield row
            else:
                curr_line += 1


def copy_dhl_template_to_export_folder(src, dst):
    """This function just uses the shutil copyfile func to copy the template
    DHL xlsx over to the export file to be ready to be edited"""
    copyfile(src, dst)


def convert_to_dhl(input_file_path, dhl_file_path):
    """Takes the shopify format and rearranges it into the dhl xlsx format"""
    line_num = 2
    for row in get_shopify_csv_by_line(input_file_path):
        add_line_into_dhl(row, dhl_file_path, line_num)
        line_num += 1


def add_line_into_dhl(row, dhl_file_path, line_num):
    """This function maps the fields from the shopify .csv into the dhl
    .xlsx
    The full mapping is not done as it does not account for orders with multiple
    items."""
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.load_workbook(dhl_file_path)

    # Get workbook active sheet
    # from the active attribute
    ws = wb.active

    # write to the export file
    pick_up_account_number = ws[xpick_up_account_number + str(line_num)]
    pick_up_account_number.value = PICKUPACCOUNTNUMBER
    shipment_order_id = ws[xshipment_order_id + str(line_num)]
    shipment_order_id.value = row[name].replace('#', '')
    consignee_name = ws[xconsignee_name + str(line_num)]
    consignee_name.value = row[shipping_name]
    address_line_1 = ws[xaddress_line_1 + str(line_num)]
    address_line_1.value = row[shipping_street]
    city = ws[xcity + str(line_num)]
    city.value = row[shipping_city]
    state = ws[xstate + str(line_num)]
    state.value = row[shipping_province]
    postal_code = ws[xpostal_code + str(line_num)]
    postal_code.value = row[shipping_zip]
    destination_country_code = ws[xdest_country_code + str(line_num)]
    destination_country_code.value = row[shipping_country]
    phone_number = ws[xphone_number + str(line_num)]
    phone_number.value = row[shipping_phone]
    email_address = ws[xemail_address + str(line_num)]
    email_address.value = row[email]
    currency_code = ws[xcurrency_code + str(line_num)]
    currency_code.value = row[currency]

    if not row[discount_amount]:
        total_declared_value = ws[xtotal_declared_value + str(line_num)]
        total_declared_value.value = row[total]
    else:
        total_declared_value = ws[xtotal_declared_value + str(line_num)]
        total_declared_value.value = float(row[total]) + float(
            row[discount_amount])

    shipment_description = ws[xshipment_description + str(line_num)]
    shipment_description.value = 'watch straps'
    content_description = ws[xcontent_description + str(line_num)]
    content_description.value = row[lineitem_name]
    content_unit_price = ws[xcontent_unit_price + str(line_num)]
    content_unit_price.value = row[lineitem_price]
    content_origin_country = ws[xcontent_origin_country + str(line_num)]
    content_origin_country.value = 'SG'
    content_quantity = ws[xcontent_quantity + str(line_num)]
    content_quantity.value = row[lineitem_quantity]
    content_weight_g = ws[xcontent_weight_g + str(line_num)]
    content_weight_g.value = check_product(row[lineitem_name])
    content_code = ws[xcontent_code + str(line_num)]
    content_code.value = row[lineitem_sku]

    # those that need to do checks before adding
    shipment_weight_g = ws[xshipment_weight_g + str(line_num)]
    shipment_weight_g.value = int(ENVELOPE) + check_product(row[lineitem_name])

    # save the changes we made to the xlsx
    wb.save(dhl_file_path)


def get_shipping_srv_code(to_country):
    """This function returns the type of shipping method based on the country
    code provided in the parameter"""
    if to_country in lst_plt_countries:
        return 'PLT'
    else:
        return 'PPS'


def check_product(default_item_in_order):
    """This function checks and returns the weight of the item that is being
    order"""
    item_in_order = default_item_in_order.lower()
    weight_of_items_in_store = {
        'sailcloth': STRAPWEIGHT,
        'deployant': DEPLOYANTWEIGHT
    }
    for key in weight_of_items_in_store:
        if key in item_in_order:
            return int(weight_of_items_in_store[key])
    return '69'


def check_multiple_orders(dhl_xlsx):
    """This function will check the Shipment Order ID column for orders with
    more than a line (multiple items in an order)"""
    # Will map into a dict
    # big_orders = {'4055': ['<row>', '<number of items>']}

    wb_obj = openpyxl.load_workbook(dhl_xlsx)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    # Loop will print all values of first column
    dict_row_occurance = {}  # map the occurrence to a dict
    for i in range(1, m_row + 1):
        # dont want the first row (header) to be shown
        if i != 1:
            cell_obj = sheet_obj.cell(row=i,
                                      column=cifs('C'))
            cell_val = str(cell_obj.value)
            if cell_val not in dict_row_occurance:
                dict_row_occurance[cell_val] = [i, 1]
            else:
                dict_row_occurance[cell_val][1] += 1
    filtered_dict = {k: v for k, v in dict_row_occurance.items() if v[1] > 1}
    populate_additional_data(wb_obj, sheet_obj, filtered_dict)


def populate_additional_data(wb, ws, filtered_dict):
    """This function will work on the multiple objects in an order
    like total weight and also the populating missing lines"""
    for k, v in filtered_dict.items():
        start_row = v[0]  # copy from this row, to every other row til end_row
        end_row = v[0] + v[1] - 1
        # ref fields to be used
        ref_consignee_name = ws[xconsignee_name + str(start_row)]
        ref_address_line_1 = ws[xaddress_line_1 + str(start_row)]
        ref_city = ws[xcity + str(start_row)]
        ref_state = ws[xstate + str(start_row)]
        ref_postal_code = ws[xpostal_code + str(start_row)]
        ref_dest_country_code = ws[xdest_country_code + str(start_row)]
        ref_phone_number = ws[xphone_number + str(start_row)]
        ref_currency_code = ws[xcurrency_code + str(start_row)]
        ref_total_declared_value = ws[xtotal_declared_value + str(start_row)]
        # /ref fields to be used
        parcel_weight_g = 0
        for i in range(start_row, end_row + 1):
            if i != start_row:
                consignee_name = ws[xconsignee_name + str(i)]
                consignee_name.value = ref_consignee_name.value
                address_line_1 = ws[xaddress_line_1 + str(i)]
                address_line_1.value = ref_address_line_1.value
                city = ws[xcity + str(i)]
                city.value = ref_city.value
                state = ws[xstate + str(i)]
                state.value = ref_state.value
                postal_code = ws[xpostal_code + str(i)]
                postal_code.value = ref_postal_code.value
                dest_country_code = ws[xdest_country_code + str(i)]
                dest_country_code.value = ref_dest_country_code.value
                phone_number = ws[xphone_number + str(i)]
                phone_number.value = ref_phone_number.value
                currency_code = ws[xcurrency_code + str(i)]
                currency_code.value = ref_currency_code.value
                total_declared_value = ws[xtotal_declared_value + str(i)]
                total_declared_value.value = ref_total_declared_value.value

                get_weight = ws[xcontent_weight_g + str(i)]
                parcel_weight_g += get_weight.value
        # adding the total weight to all of the same order
        for i in range(start_row, end_row + 1):
            total_parcel_weight = ws[xtotal_declared_value + str(i)]
            total_parcel_weight.value = int(ENVELOPE) + parcel_weight_g

    # add shipment service code
    m_row = ws.max_row
    for i in range(2, m_row + 1):
        # now its time to add in the Shipment Service Code (PPS, PLT)
        to_country = ws.cell(row=i, column=cifs(xdest_country_code))
        shipping_srv_code = ws.cell(row=i, column=cifs(xshipping_service_code))
        shipping_srv_code.value = get_shipping_srv_code(to_country.value)

        # print(shipping_srv_code.value)
        if shipping_srv_code.value in dict_incoterm.keys():
            incoterm = ws.cell(row=i, column=cifs(xincoterm))
            incoterm.value = dict_incoterm[shipping_srv_code.value]

    wb.save(EXPORT_PATH_NAME.format(EXPORT_NAME))


def main(optional_param=''):
    """This function will serve as a starting point for the program"""
    try:
        if optional_param == '':
            shopify_csv = sys.argv[1]
        else:
            shopify_csv = optional_param

        # cp the template into the export folder for editing
        copy_dhl_template_to_export_folder(DHL_TEMPLATE_FILE, EXPORT_PATH_NAME)

        # start the conversion process
        convert_to_dhl(shopify_csv, EXPORT_PATH_NAME)

        # check for multiple items in an order
        check_multiple_orders(EXPORT_PATH_NAME)

    except IndexError:
        print('Please enter Shopify order csv as parameter.')
    except UnicodeDecodeError:
        pass


if __name__ == '__main__':
    main()
